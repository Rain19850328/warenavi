# app/main.py
# -*- coding: utf-8 -*-
import base64
import datetime
import io
import json
import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from threading import Lock
from typing import Dict, List, Tuple, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import bindparam, create_engine, text
from sqlalchemy.exc import SQLAlchemyError

def model_to_dict(m):
  # pydantic v2 우선, v1 호환
  return m.model_dump() if hasattr(m, "model_dump") else m.dict()

DB_URL = os.environ.get("ITEMS_DB_URL", "mysql+pymysql://rain:8520@192.168.0.5:3306/items")
MOPS_DB_URL = os.environ.get("MOPS_DB_URL", "mysql+pymysql://rain:8520@192.168.0.5:3306/mops_2026")
NEW_INBOUND_STORE_PATH = Path(__file__).resolve().parent / "data" / "new_inbound_lists.json"
SALES_DB_URL = os.environ.get("SALES_DB_URL", "mysql+pymysql://rain:8520@192.168.0.5:3306/salesdb")  # ★ 추가


ROW_IDS = [f"SR{i}" for i in range(1, 4)]
LEVEL_WIDTH_CM  = 110
LEVEL_DEPTH_CM  = 110
# SR-02가 6레벨을 쓰므로 4~6 기본 높이 추가(임의 100)
LEVEL_HEIGHT_CM = {1:120, 2:100, 3:180, 4:100, 5:100, 6:100}
LEVEL_CAPACITY_CM3 = {lv: LEVEL_WIDTH_CM * LEVEL_DEPTH_CM * LEVEL_HEIGHT_CM[lv] for lv in LEVEL_HEIGHT_CM}
DEFAULT_BAYS, DEFAULT_LEVELS = 21, 3
RATE_LOW_MAX, RATE_NORMAL_MAX = 25, 70

# 행별 레이아웃(베이 수, 레벨 수) — SR-02: 11열, 6레벨
LAYOUT_BY_ROW = {"SR1": (21, 3), "SR2": (11, 6), "SR3": (21, 3)}

# ---------------- SR-02 문자 레벨(A~F) 지원 유틸 ----------------
SR2_LETTER_CODE_RE = re.compile(r"^SR2-([A-Fa-f])-([0-9]{2})$")
SR2_NUM_CODE_RE    = re.compile(r"^SR2-([0-9]{2})-([0-9]{2})$")

def sr2_level_to_letter(level: int) -> str:
  max_lv = LAYOUT_BY_ROW.get("SR2", (11, 6))[1]
  return chr(65 + (max_lv - int(level)))  # level=6 -> 'A'

def sr2_letter_to_level(letter: str) -> int:
  max_lv = LAYOUT_BY_ROW.get("SR2", (11, 6))[1]
  idx = ord(letter.upper()) - 65
  if idx < 0 or idx >= max_lv:
    raise ValueError("잘못된 SR2 레벨 문자")
  return max_lv - idx  # 'A'->6, 'F'->1

def parse_rack_code(code: str) -> Optional[Tuple[str,int,int]]:
  s = (code or "").strip()

  # 신규 포맷: SR2-A-01 / SR2-01-06
  m = SR2_LETTER_CODE_RE.match(s)
  if m:
    letter, bay_s = m.groups()
    return ("SR2", int(bay_s), sr2_letter_to_level(letter))

  m = SR2_NUM_CODE_RE.match(s)  # 혹시 숫자 레벨이 들어오면 파싱
  if m:
    bay_s, level_s = m.groups()
    return ("SR2", int(bay_s), int(level_s))
  
  # ✅ 새 포맷(숫자 레벨, SR1/SR2/SR3…) 지원: SR1-05-02, SR3-10-01 등
  m = re.match(r"^SR(\d{1,2})-(\d{2})-(\d{2})$", s)
  if m:
    row_s, bay_s, level_s = m.groups()
    return (f"SR{int(row_s)}", int(bay_s), int(level_s))

  # 구형 포맷 호환: SR-02-A-01 / SR-02-01-06 / SR-01-.. 등
  m = re.match(r"^SR-(\d{2})-([A-Fa-f])-([0-9]{2})$", s)
  if m and m.group(1) == "02":
    _, letter, bay_s = m.groups()
    return ("SR2", int(bay_s), sr2_letter_to_level(letter))

  m = re.match(r"^SR-(\d{2})-(\d{2})-(\d{2})$", s)
  if m:
    row_s, bay_s, level_s = m.groups()
    return (f"SR{int(row_s)}", int(bay_s), int(level_s))

  return None

def sr02_numeric_to_letter(code: str) -> str:
  """SR2-01-06 또는 SR-02-01-06 -> SR2-A-01 로 변환(검색 편의)"""
  s = (code or "").strip()
  m = re.match(r'^(?:SR-?0?2|SR2)-([0-9]{2})-([0-9]{2})$', s)
  if not m:
    return code
  bay_s, level_s = m.groups()
  letter = sr2_level_to_letter(int(level_s))
  return f"SR2-{letter}-{int(bay_s):02d}"

def to_canonical_code(code: str) -> str:
  """입력 형태가 무엇이든 DB 저장용 표준 코드로 변환.
     SR2는 문자레벨(A~F), 그 외는 숫자레벨 유지."""
  p = parse_rack_code(code)
  if not p:
    return (code or "").strip()
  row, bay, level = p
  if row == "SR2":
    return f"{row}-{sr2_level_to_letter(level)}-{bay:02d}"  # 예: SR2-A-01
  return f"{row}-{bay:02d}-{level:02d}"                     # 예: SR1-05-02 

def _has_wildcard_00(loc: str) -> bool:
  """로케이션 코드에 '00' 세그먼트가 있으면(예: SR1-00-02, SR1-05-00) True"""
  s = (loc or "").strip()
  # 캐논형으로 맞춘 뒤, 하이픈 단위로 '00'이 포함되면 와일드카드로 간주
  can = to_canonical_code(s)
  parts = can.split("-")
  return any(p == "00" for p in parts)

# ---------------------------------------------------------------

def occupancy_rate(used_cm3: int, cap_cm3: int) -> float:
  return 0.0 if cap_cm3 <= 0 else (used_cm3 / cap_cm3) * 100.0

@dataclass
class CellInfo:
  code: str
  capacity: int
  total_cm3: int
  top_sku: Optional[str]
  kinds: int
  items: List[Tuple[str, str, int, int, str]]  # (sku, name, qty, used_cm3, location)

class DataProvider:
  def rows(self) -> List[str]: raise NotImplementedError
  def layout(self) -> Tuple[int,int]: return (DEFAULT_BAYS, DEFAULT_LEVELS)
  # 행별 레이아웃
  def layout_for_row(self, row_label: str) -> Tuple[int,int]:
    return LAYOUT_BY_ROW.get(row_label, (DEFAULT_BAYS, DEFAULT_LEVELS))
  def fetch_cells_for_row(self, row_label: str) -> Dict[Tuple[int,int], CellInfo]: raise NotImplementedError
  def recent_movements(self, limit:int=30) -> List[Tuple[str,str,str,int,str]]: return []
  def list_items(self, query: str="", limit:int=500) -> List[Tuple[str,str,float,float,float,str]]: raise NotImplementedError
  def inbound(self, rack_code: str, item_code: str, qty: int) -> None: raise NotImplementedError
  def outbound(self, rack_code: str, item_code: str, qty: int) -> None: raise NotImplementedError
  def move(self, from_rack: str, to_rack: str, item_code: str, qty: int) -> None: raise NotImplementedError
  def search_racks(self, query: str="", limit:int=500) -> List[dict]: raise NotImplementedError
  def location_in_use(self, location: str) -> Optional[str]:
    """해당 location을 사용 중인 item_code가 있으면 그 코드를 반환, 없으면 None."""
    return None
  def set_location(self, item_code: str, location: str) -> None:
    """아이템의 location 코드를 저장(중복이면 예외)."""
    raise NotImplementedError

class MySQLProvider(DataProvider):
  # 기존 LOC_RE는 호환용(다른 행의 숫자코드 파싱에만 사용 가능) — SR-02는 parse_rack_code 사용
  LOC_RE = re.compile(r"^SR-(\d{2})-(\d{2})-(\d{2})$")
  def __init__(self, db_url: str):
    self.engine = create_engine(
        db_url,
        pool_pre_ping=True,
        future=True,
        pool_recycle=1800,
        isolation_level="READ COMMITTED",  # 선택
    )
    self.sales_engine = create_engine(
        SALES_DB_URL,
        pool_pre_ping=True,
        future=True,
        pool_recycle=1800,
        isolation_level="READ COMMITTED",
    )
    self._has_items = self._check_items_table()

    self._schema_mode = self._detect_schema_rack()
    self._item_loc_col = self._detect_items_location_col()
    self._data: Dict[str, Dict[Tuple[int,int], CellInfo]] = {}
    self._load()

  def _check_items_table(self) -> bool:
    try:
      with self.engine.connect() as conn:
        sql = text("""
          SELECT COUNT(*) FROM information_schema.tables
          WHERE table_schema = DATABASE() AND table_name = 'items'
        """)
        return bool(conn.execute(sql).scalar())
    except Exception:
      return False

  def rows(self) -> List[str]: return ROW_IDS

  def layout_for_row(self, row_label: str) -> Tuple[int,int]:
    return LAYOUT_BY_ROW.get(row_label, (DEFAULT_BAYS, DEFAULT_LEVELS))

  def _detect_schema_rack(self) -> str:
    try:
      with self.engine.connect() as conn:
        rs = list(conn.execute(text("SHOW COLUMNS FROM storage_racks LIKE 'item_id'")))
        if rs: return "id"
        rs2 = list(conn.execute(text("SHOW COLUMNS FROM storage_racks LIKE 'item_code'")))
        return "code" if rs2 else "code"
    except Exception:
      return "code"

  def _detect_items_location_col(self) -> Optional[str]:
    if not getattr(self, "_has_items", False):
      return None
    try:
      with self.engine.connect() as conn:
        rs = list(conn.execute(text("SHOW COLUMNS FROM items LIKE 'location_code'")))
        if rs: return "location_code"
        rs2 = list(conn.execute(text("SHOW COLUMNS FROM items LIKE 'locationcode'")))
        if rs2: return "locationcode"
    except Exception:
      pass
    return None

  def list_items(self, query: str="", limit:int=500) -> List[Tuple[str,str,float,float,float,str]]:
    loc_col = self._item_loc_col
    select_loc = f", COALESCE(i.{loc_col}, '') AS loc" if loc_col else ", '' AS loc"
    where_loc  = f" OR i.{loc_col} LIKE :pat" if loc_col else ""
    sql = text(f"""
      SELECT i.code, i.name,
             COALESCE(i.volume_width,0.0)  AS w,
             COALESCE(i.volume_length,0.0) AS l,
             COALESCE(i.volume_height,0.0) AS h
             {select_loc}
      FROM items i
      WHERE (:q = '' OR i.code LIKE :pat OR i.name LIKE :pat{where_loc})
      ORDER BY i.code ASC, i.name ASC
      LIMIT :lim
    """)
    q = (query or "").strip()
    pat = f"%{q}%"
    with self.engine.connect() as conn:
      rows = list(conn.execute(sql, {"q":q, "pat":pat, "lim":limit}))
    return [(r[0], r[1], float(r[2]), float(r[3]), float(r[4]), r[5] or "") for r in rows]

  def inbound(self, rack_code: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    mode = self._schema_mode
    try:
      with self.engine.begin() as conn:
        if mode == "id":
          rid = conn.execute(text("SELECT id FROM items WHERE code = :c"), {"c": item_code}).scalar()
          if not rid: raise RuntimeError(f"SKU를 찾을 수 없습니다: {item_code}")
          conn.execute(text("""
            INSERT INTO storage_racks (rack_code, item_id, quantity)
            VALUES (:rack, :item_id, :qty)
            ON DUPLICATE KEY UPDATE quantity = quantity + :qty_u
          """), {"rack": rack_code, "item_id": int(rid), "qty": int(qty), "qty_u": int(qty)})
        else:
          conn.execute(text("""
            INSERT INTO storage_racks (rack_code, item_code, quantity)
            VALUES (:rack, :item_code, :qty)
            ON DUPLICATE KEY UPDATE quantity = quantity + :qty_u
          """), {"rack": rack_code, "item_code": item_code, "qty": int(qty), "qty_u": int(qty)})
    except SQLAlchemyError as e:
      raise RuntimeError(f"입고 처리 실패: {e}")
    self._load()

  def _get_current_qty(self, conn, rack_code: str, item_code: str, mode: str) -> int:
    if mode == "id":
      rid = conn.execute(text("SELECT id FROM items WHERE code=:c"), {"c": item_code}).scalar()
      if not rid: return 0
      return int(conn.execute(text("""
        SELECT COALESCE(quantity,0) FROM storage_racks WHERE rack_code=:rack AND item_id=:rid
      """), {"rack": rack_code, "rid": int(rid)}).scalar() or 0)
    else:
      return int(conn.execute(text("""
        SELECT COALESCE(quantity,0) FROM storage_racks WHERE rack_code=:rack AND item_code=:code
      """), {"rack": rack_code, "code": item_code}).scalar() or 0)

  def outbound(self, rack_code: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    mode = self._schema_mode
    try:
      with self.engine.begin() as conn:
        cur = self._get_current_qty(conn, rack_code, item_code, mode)
        if cur < qty: raise RuntimeError("수량 부족")
        if mode == "id":
          rid = conn.execute(text("SELECT id FROM items WHERE code=:c"), {"c": item_code}).scalar()
          conn.execute(text("""
            UPDATE storage_racks SET quantity = quantity - :qty
            WHERE rack_code=:rack AND item_id=:rid
          """), {"rack": rack_code, "rid": int(rid), "qty": int(qty)})
          conn.execute(text("""
            DELETE FROM storage_racks WHERE rack_code=:rack AND item_id=:rid AND quantity<=0
          """), {"rack": rack_code, "rid": int(rid)})
        else:
          conn.execute(text("""
            UPDATE storage_racks SET quantity = quantity - :qty
            WHERE rack_code=:rack AND item_code=:code
          """), {"rack": rack_code, "code": item_code, "qty": int(qty)})
          conn.execute(text("""
            DELETE FROM storage_racks WHERE rack_code=:rack AND item_code=:code AND quantity<=0
          """), {"rack": rack_code, "code": item_code})
    except SQLAlchemyError as e:
      raise RuntimeError(f"출고 처리 실패: {e}")
    self._load()

  def move(self, from_rack: str, to_rack: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    if from_rack == to_rack: return
    mode = self._schema_mode
    try:
      with self.engine.begin() as conn:
        cur = self._get_current_qty(conn, from_rack, item_code, mode)
        if cur < qty: raise RuntimeError("수량 부족")
        if mode == "id":
          rid = conn.execute(text("SELECT id FROM items WHERE code=:c"), {"c": item_code}).scalar()
          conn.execute(text("""
            UPDATE storage_racks SET quantity = quantity - :qty
            WHERE rack_code=:rack AND item_id=:rid
          """), {"rack": from_rack, "rid": int(rid), "qty": int(qty)})
          conn.execute(text("""
            DELETE FROM storage_racks WHERE rack_code=:rack AND item_id=:rid AND quantity<=0
          """), {"rack": from_rack, "rid": int(rid)})
          conn.execute(text("""
            INSERT INTO storage_racks (rack_code, item_id, quantity)
            VALUES (:rack, :rid, :qty)
            ON DUPLICATE KEY UPDATE quantity = quantity + :qty_u
          """), {"rack": to_rack, "rid": int(rid), "qty": int(qty), "qty_u": int(qty)})
        else:
          conn.execute(text("""
            UPDATE storage_racks SET quantity = quantity - :qty
            WHERE rack_code=:rack AND item_code=:code
          """), {"rack": from_rack, "code": item_code, "qty": int(qty)})
          conn.execute(text("""
            DELETE FROM storage_racks WHERE rack_code=:rack AND item_code=:code AND quantity<=0
          """), {"rack": from_rack, "code": item_code})
          conn.execute(text("""
            INSERT INTO storage_racks (rack_code, item_code, quantity)
            VALUES (:rack, :code, :qty)
            ON DUPLICATE KEY UPDATE quantity = quantity + :qty_u
          """), {"rack": to_rack, "code": item_code, "qty": int(qty), "qty_u": int(qty)})
    except SQLAlchemyError as e:
      raise RuntimeError(f"이동 처리 실패: {e}")
    self._load()

  # ★★★ 신규: 최신 일자 기준 stock_cnt_real 일괄 조회
  def latest_stock_for_codes(self, codes: List[str]) -> Dict[str, int]:
    if not codes: return {}
    eng = getattr(self, "sales_engine", None)
    if not eng: return {}
    # MySQL8: 서브쿼리 조인으로 최신 date 매칭
    sql = text("""
      SELECT s.sku_cd, s.stock_cnt_real
      FROM daily_stock s
      JOIN (
        SELECT sku_cd, MAX(date) AS maxd
        FROM daily_stock
        GROUP BY sku_cd
      ) t ON t.sku_cd = s.sku_cd AND t.maxd = s.date
      WHERE s.sku_cd IN :codes
    """)
    # SQLAlchemy IN 바인딩은 tuple 권장
    params = {"codes": tuple(codes)}
    out = {}
    try:
      with eng.connect() as conn:
        for r in conn.execute(sql, params):
          # sku_cd가 TEXT 이므로 공백 트리밍
          key = (r[0] or "").strip()
          if key:
            out[key] = int(r[1] or 0)
    except Exception:
      return {}
    return out

  def _try_query(self, by: str):
    has_items = getattr(self, "_has_items", False)
    loc_col = self._item_loc_col if has_items else None
    select_loc = f", COALESCE(i.{loc_col}, '') AS loc" if loc_col else ", '' AS loc"

    if has_items:
      # LEFT JOIN 사용 (일부 코드가 items에 없어도 동작)
      on = "i.id = sr.item_id" if by == "id" else "i.code = sr.item_code"
      sql = text(f"""
        SELECT sr.rack_code AS rack, COALESCE(sr.quantity, 0) AS qty,
              COALESCE(i.code, sr.item_code) AS sku,
              COALESCE(i.name, '') AS nm,
              COALESCE(i.volume_width,0.0)  AS w,
              COALESCE(i.volume_length,0.0) AS l,
              COALESCE(i.volume_height,0.0) AS h
              {select_loc}
        FROM storage_racks sr
        LEFT JOIN items i ON {on}
      """)
    else:
      # items 테이블이 없을 때: storage_racks만으로 구성
      sql = text("""
        SELECT sr.rack_code AS rack, COALESCE(sr.quantity,0) AS qty,
              sr.item_code AS sku, '' AS nm,
              0.0 AS w, 0.0 AS l, 0.0 AS h, '' AS loc
        FROM storage_racks sr
      """)

    with self.engine.connect() as conn:
      return list(conn.execute(sql).mappings())

  def _load(self):
    try:
      rows = self._try_query("id")
    except Exception:
      rows = []
    if not rows:
      try:
        rows = self._try_query("code")
      except Exception as e:
        raise RuntimeError(f"storage_racks JOIN 실패: {e}")

    used_cm3: Dict[str, Dict[Tuple[int,int], float]] = {}
    vol_by_sku: Dict[str, Dict[Tuple[int,int], Dict[str,float]]] = {}
    qty_by_sku: Dict[str, Dict[Tuple[int,int], Dict[str,int]]] = {}
    name_by_sku: Dict[str, Dict[Tuple[int,int], Dict[str,str]]] = {}
    loc_by_sku: Dict[str, Dict[Tuple[int,int], Dict[str,str]]] = {}

    for r in rows:
      rack = str(r["rack"] or "").strip()
      parsed = parse_rack_code(rack)
      if not parsed:
        continue
      row_label, bay, level = parsed

      bays_row, levels_row = self.layout_for_row(row_label)
      if not (1 <= bay <= bays_row and 1 <= level <= levels_row):
        continue

      sku = (r["sku"] or "").strip()
      nm  = (r["nm"]  or "").strip()
      w   = float(r["w"] or 0.0); l = float(r["l"] or 0.0); h = float(r["h"] or 0.0)
      qty = int(r["qty"] or 0)
      loc = str(r.get("loc") or "")

      unit_cm3 = max(0.0, w) * max(0.0, l) * max(0.0, h)
      used = unit_cm3 * max(0, qty)

      key = (bay, level)
      used_cm3.setdefault(row_label, {})
      used_cm3[row_label][key] = used_cm3[row_label].get(key, 0.0) + used

      vol_by_sku.setdefault(row_label, {}).setdefault(key, {})
      vol_by_sku[row_label][key][sku] = vol_by_sku[row_label][key].get(sku, 0.0) + used

      qty_by_sku.setdefault(row_label, {}).setdefault(key, {})
      qty_by_sku[row_label][key][sku] = qty_by_sku[row_label][key].get(sku, 0) + qty

      name_by_sku.setdefault(row_label, {}).setdefault(key, {})
      name_by_sku[row_label][key][sku] = nm

      loc_by_sku.setdefault(row_label, {}).setdefault(key, {})
      loc_by_sku[row_label][key][sku] = loc

    self._data = {}
    for row_label in ROW_IDS:
      self._data[row_label] = {}
      bays_row, levels_row = self.layout_for_row(row_label)
      for bay in range(1, bays_row+1):
        for level in range(1, levels_row+1):
          key = (bay, level)
          total = used_cm3.get(row_label, {}).get(key, 0.0)
          sku_map = vol_by_sku.get(row_label, {}).get(key, {})
          items_sorted = sorted(sku_map.items(), key=lambda x: x[1], reverse=True)
          kinds = len(items_sorted)
          top_sku = items_sorted[0][0] if items_sorted else None
          cap = int(LEVEL_CAPACITY_CM3.get(level, LEVEL_CAPACITY_CM3[1]))

          items_list: List[Tuple[str, str, int, int, str]] = []
          for sku, vol in items_sorted:
            nm = name_by_sku.get(row_label, {}).get(key, {}).get(sku, "")
            q  = qty_by_sku.get(row_label, {}).get(key, {}).get(sku, 0)
            lc = loc_by_sku.get(row_label, {}).get(key, {}).get(sku, "")
            items_list.append((sku, nm, q, int(vol), lc))

          # 가시 코드: SR-02는 문자 레벨로 노출 (SR-02-A-01)
          code = f"{row_label}-{bay:02d}-{level:02d}"
          if row_label == "SR2":
            letter = sr2_level_to_letter(level)
            code = f"{row_label}-{letter}-{bay:02d}"

          self._data[row_label][key] = CellInfo(
            code=code,
            capacity=int(cap),
            total_cm3=int(total),
            top_sku=top_sku,
            kinds=kinds,
            items=items_list,
          )

  def fetch_cells_for_row(self, row_label: str) -> Dict[Tuple[int,int], CellInfo]:
    return self._data.get(row_label, {})

  def recent_movements(self, limit:int=30) -> List[Tuple[str,str,str,int,str]]:
    now = datetime.datetime.now()
    out = []
    for i in range(min(limit, 20)):
      out.append(((now - datetime.timedelta(minutes=i*7)).strftime("%H:%M"),
                  ["입고","출고","이동","조정"][i%4],
                  "SKU-EXAMPLE",
                  [+30, +50, -10, -4][i%4],
                  "—"))
    return out

  def search_racks(self, query: str="", limit:int=500) -> List[dict]:
    q = (query or "").strip()

    # 1차 패턴: 원문
    pat = f"%{q}%"

    # 2차 패턴: 파싱해 캐논컬 코드로 변환(SR2는 문자레벨 강제)
    canon = None
    parsed = parse_rack_code(q)
    if parsed:
      row_c, bay_c, level_c = parsed
      if row_c == "SR2":
        canon = f"SR2-{sr2_level_to_letter(level_c)}-{bay_c:02d}"
      else:
        canon = f"{row_c}-{bay_c:02d}-{level_c:02d}"
    pat2 = f"%{canon}%" if canon else pat
    mode = self._schema_mode
    loc_col = getattr(self, "_item_loc_col", None)
    select_loc = f", COALESCE(i.{loc_col}, '') AS loc" if loc_col else ", '' AS loc"
    where_loc  = f" OR i.{loc_col} LIKE :pat" if loc_col else ""

    if mode == "id":
      sql = text(f"""
        SELECT sr.rack_code AS rack, COALESCE(sr.quantity,0) AS qty,
               i.code AS sku, i.name AS nm,
               COALESCE(i.volume_width,0.0)  AS w,
               COALESCE(i.volume_length,0.0) AS l,
               COALESCE(i.volume_height,0.0) AS h
               {select_loc}
        FROM storage_racks sr
        JOIN items i ON i.id = sr.item_id
        WHERE (:q = '' OR i.code LIKE :pat OR i.name LIKE :pat{where_loc}
              OR sr.rack_code LIKE :pat OR sr.rack_code LIKE :pat2)
        ORDER BY sr.rack_code ASC, i.code ASC
        LIMIT :lim
      """)
    else:
      sql = text(f"""
        SELECT sr.rack_code AS rack, COALESCE(sr.quantity,0) AS qty,
               i.code AS sku, i.name AS nm,
               COALESCE(i.volume_width,0.0)  AS w,
               COALESCE(i.volume_length,0.0) AS l,
               COALESCE(i.volume_height,0.0) AS h
               {select_loc}
        FROM storage_racks sr
        JOIN items i ON i.code = sr.item_code
        WHERE (:q = '' OR i.code LIKE :pat OR i.name LIKE :pat{where_loc}
               OR sr.rack_code LIKE :pat OR sr.rack_code LIKE :pat2)
        ORDER BY sr.rack_code ASC, i.code ASC
        LIMIT :lim
      """)
    with self.engine.connect() as conn:
      rows = list(conn.execute(sql, {"q": q, "pat": pat, "pat2": pat2, "lim": int(limit)}).mappings())

    out: List[dict] = []
    for r in rows:
      rack = str(r["rack"] or "").strip()
      parsed = parse_rack_code(rack)
      if not parsed:
        continue
      row_label, bay, level = parsed
      sku = (r["sku"] or "").strip()
      nm  = (r["nm"]  or "").strip()
      qty = int(r["qty"] or 0)
      w   = float(r["w"] or 0.0); l = float(r["l"] or 0.0); h = float(r["h"] or 0.0)
      unit = max(0.0,w)*max(0.0,l)*max(0.0,h)
      used = int(unit * max(0, qty))
      cap  = int(LEVEL_CAPACITY_CM3.get(level, LEVEL_CAPACITY_CM3[1]))
      rate = int(occupancy_rate(used, cap))

      # 결과 노출 코드는 SR-02면 문자 레벨 표기 강제
      rack_vis = rack
      if row_label == "SR2":
        letter = sr2_level_to_letter(level)
        rack_vis = f"SR2-{letter}-{bay:02d}"

      out.append({
        "rack_code": rack_vis,
        "row": row_label,
        "bay": bay,
        "level": level,
        "sku": sku,
        "name": nm,
        "qty": qty,
        "capacity_cm3": cap,
        "used_cm3": used,
        "rate": rate,
        "location": (r.get("loc") or ""),
      })
    return out

  def location_in_use(self, location: str) -> Optional[str]:
    # '00'이 들어간 위치코드는 중복 체크 무시
    if _has_wildcard_00(location):
      return None

    loc_col = getattr(self, "_item_loc_col", None)
    if not loc_col:
      return None
    sql = text(f"SELECT code FROM items WHERE {loc_col} = :loc LIMIT 1")
    with self.engine.connect() as conn:
      row = conn.execute(sql, {"loc": to_canonical_code(location)}).first()
      return (row[0] if row else None)


  def set_location(self, item_code: str, location: str) -> None:
    loc_col = getattr(self, "_item_loc_col", None)
    if not self._has_items or not loc_col:
      raise RuntimeError("items 테이블 또는 위치 컬럼이 없습니다.")

    # SR2 등은 캐논컬 형태로 저장
    loc_norm = to_canonical_code(location)

    # 중복 체크
    used_by = self.location_in_use(loc_norm)
    if used_by and used_by != item_code:
      raise RuntimeError("해당 로케이션 코드는 이미 다른 상품이 사용 중입니다.")

    with self.engine.begin() as conn:
      r = conn.execute(
        text(f"UPDATE items SET {loc_col} = :loc WHERE code = :code"),
        {"loc": loc_norm, "code": item_code},
      )
      if r.rowcount == 0:
        raise RuntimeError("아이템 코드를 찾을 수 없습니다.")

    # 캐시 갱신
    self._load()  

class FakeProvider(DataProvider):
  def __init__(self):
    self._rows = ROW_IDS[:]
    self._cells: Dict[str, Dict[Tuple[int,int], CellInfo]] = {r:{} for r in self._rows}
    self._items = [
      ("A00100301", "샘플상품A(14×18.5×3)", 14.0, 18.5, 3.0, "C-03-03-01"),
      ("B00000001", "샘플상품B(20×20×20)", 20.0, 20.0, 20.0, "SR1-05-02"),
      ("C00000001", "샘플상품C(30×25×15)", 30.0, 25.0, 15.0, ""),
    ]
    for r in self._rows:
      bays_row, levels_row = self.layout_for_row(r)
      for b in range(1, bays_row+1):
        for lv in range(1, levels_row+1):
          code = f"{r}-{b:02d}-{lv:02d}"
          if r == "SR2":
            code = f"{r}-{sr2_level_to_letter(lv)}-{b:02d}"
          self._cells[r][(b,lv)] = CellInfo(
            code=code,
            capacity=int(LEVEL_CAPACITY_CM3.get(lv, LEVEL_CAPACITY_CM3[1])),
            total_cm3=0, top_sku=None, kinds=0, items=[]
          )
    used = int(14.0 * 18.5 * 3.0 * 2000)
    self._cells["SR3"][(3,3)] = CellInfo(
      code="SR3-03-03",
      capacity=LEVEL_CAPACITY_CM3[3],
      total_cm3=used, top_sku="A00100301", kinds=1,
      items=[("A00100301", "샘플상품A(14×18.5×3)", 2000, used, "C-03-03-01")]
    )

  def rows(self) -> List[str]: return self._rows

  def layout_for_row(self, row_label: str) -> Tuple[int,int]:
    return LAYOUT_BY_ROW.get(row_label, (DEFAULT_BAYS, DEFAULT_LEVELS))

  def list_items(self, query: str="", limit:int=500) -> List[Tuple[str,str,float,float,float,str]]:
    q = (query or "").strip().lower()
    out = []
    for code, name, w, l, h, loc in sorted(self._items, key=lambda x: (x[1].lower(), x[0].lower())):
      if not q or q in code.lower() or q in name.lower() or q in (loc or "").lower():
        out.append((code, name, w, l, h, loc or ""))
        if len(out) >= limit: break
    return out

  def _apply_delta(self, rack_code: str, item_code: str, delta_qty: int):
    parsed = parse_rack_code(rack_code)
    if not parsed: raise ValueError("위치 코드 형식 오류")
    row, bay, level = parsed
    matches = [it for it in self._items if it[0] == item_code]
    if not matches: raise RuntimeError("아이템을 찾을 수 없습니다.")
    code, name, w, l, h, _loc = matches[0]
    unit_cm3 = w*l*h
    used_delta = int(unit_cm3 * abs(delta_qty))
    ci = self._cells[row][(bay,level)]
    idx = None
    for i, (sku, nm, q, vol, loc) in enumerate(ci.items):
      if sku == code:
        idx = i; break
    if delta_qty > 0:
      if idx is None:
        ci.items.append((code, name, delta_qty, used_delta, _loc or ""))
      else:
        sku, nm, q, vol, loc = ci.items[idx]
        ci.items[idx] = (sku, nm, q + delta_qty, vol + used_delta, loc)
      ci.total_cm3 += used_delta
    else:
      if idx is None: raise RuntimeError("재고가 없습니다.")
      sku, nm, q, vol, loc = ci.items[idx]
      if q + delta_qty < 0: raise RuntimeError("수량 부족")
      new_q = q + delta_qty
      new_vol = int(unit_cm3 * new_q)
      if new_q == 0:
        ci.total_cm3 -= vol
        del ci.items[idx]
      else:
        ci.total_cm3 += (new_vol - vol)
        ci.items[idx] = (sku, nm, new_q, new_vol, loc)
    ci.items.sort(key=lambda x: x[3], reverse=True)
    ci.kinds = len(ci.items)
    ci.top_sku = ci.items[0][0] if ci.items else None

  def inbound(self, rack_code: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    self._apply_delta(rack_code, item_code, +qty)

  def outbound(self, rack_code: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    self._apply_delta(rack_code, item_code, -qty)

  def move(self, from_rack: str, to_rack: str, item_code: str, qty: int) -> None:
    if qty <= 0: raise ValueError("수량은 1 이상이어야 합니다.")
    if from_rack == to_rack: return
    self._apply_delta(from_rack, item_code, -qty)
    self._apply_delta(to_rack, item_code, +qty)

  def fetch_cells_for_row(self, row_label: str) -> Dict[Tuple[int,int], CellInfo]:
    return self._cells.get(row_label, {})

  def recent_movements(self, limit:int=30) -> List[Tuple[str,str,str,int,str]]:
    now = datetime.datetime.now()
    return [((now - datetime.timedelta(minutes=i*5)).strftime("%H:%M"), "입고", "A00100301", +50, "—")
            for i in range(1, min(20, limit)+1)]

  def latest_stock_for_codes(self, codes: List[str]) -> Dict[str, int]:
    return {}
  def search_racks(self, query: str="", limit:int=500) -> List[dict]:
    q = (query or "").strip().lower()
    out: List[dict] = []
    for row_label, cols in self._cells.items():
      for (bay, level), ci in cols.items():
        if q and q not in ci.code.lower():
          pass
        for sku, name, qty, used, loc in ci.items:
          if not q or q in sku.lower() or q in (name or "").lower() or q in (loc or "").lower() or q in ci.code.lower():
            cap = int(LEVEL_CAPACITY_CM3.get(level, LEVEL_CAPACITY_CM3[1]))
            out.append({
              "rack_code": ci.code,
              "row": row_label,
              "bay": bay,
              "level": level,
              "sku": sku,
              "name": name,
              "qty": int(qty),
              "capacity_cm3": cap,
              "used_cm3": int(used),
              "rate": int(occupancy_rate(used, cap)),
              "location": loc or "",
            })
            if len(out) >= limit: return out
    return out

  def location_in_use(self, location: str) -> Optional[str]:
    # '00'이 들어간 위치코드는 중복 체크 무시
    if _has_wildcard_00(location):
      return None

    loc_norm = to_canonical_code(location)
    for code, _name, _w, _l, _h, loc in self._items:
      if (loc or "") == loc_norm:
        return code
    return None


  def set_location(self, item_code: str, location: str) -> None:
    loc_norm = to_canonical_code(location)
    used_by = self.location_in_use(loc_norm)
    if used_by and used_by != item_code:
      raise RuntimeError("해당 로케이션 코드는 이미 다른 상품이 사용 중입니다.")
    # 업데이트
    for i, (code, name, w, l, h, loc) in enumerate(self._items):
      if code == item_code:
        self._items[i] = (code, name, w, l, h, loc_norm)
        return
    raise RuntimeError("아이템 코드를 찾을 수 없습니다.")  

class NewInboundStore:
  def __init__(self, path: Path):
    self.path = path
    self._lock = Lock()

  def _default_payload(self) -> dict:
    return {"days": {}}

  def _read(self) -> dict:
    if not self.path.exists():
      return self._default_payload()
    try:
      return json.loads(self.path.read_text(encoding="utf-8"))
    except Exception:
      return self._default_payload()

  def _write(self, payload: dict) -> None:
    self.path.parent.mkdir(parents=True, exist_ok=True)
    self.path.write_text(
      json.dumps(payload, ensure_ascii=False, indent=2),
      encoding="utf-8",
    )

  def get_day(self, date_text: str) -> dict:
    with self._lock:
      payload = self._read()
      day = payload.get("days", {}).get(date_text)
      if not day:
        return {"date": date_text, "updated_at": "", "items": []}
      return json.loads(json.dumps(day, ensure_ascii=False))

  def replace_day(self, date_text: str, items: List[dict], source_name: str = "") -> dict:
    with self._lock:
      payload = self._read()
      payload.setdefault("days", {})
      day = {
        "date": date_text,
        "updated_at": datetime.datetime.now().isoformat(),
        "source_name": source_name or "",
        "items": items,
      }
      payload["days"][date_text] = day
      self._write(payload)
      return json.loads(json.dumps(day, ensure_ascii=False))

  def process_item(self, date_text: str, entry_id: str, qty: int, action: str, rack_code: str = "") -> dict:
    with self._lock:
      payload = self._read()
      day = payload.setdefault("days", {}).get(date_text)
      if not day:
        raise RuntimeError("해당 날짜의 신규입고리스트가 없습니다.")

      found = None
      for item in day.get("items", []):
        if str(item.get("id") or "") == str(entry_id):
          found = item
          break
      if not found:
        raise RuntimeError("선택한 항목을 찾을 수 없습니다.")

      pending_qty = int(found.get("pending_qty") or 0)
      if qty <= 0:
        raise RuntimeError("수량은 1 이상이어야 합니다.")
      if qty > pending_qty:
        raise RuntimeError("미처리수량보다 큰 수량은 처리할 수 없습니다.")

      found["pending_qty"] = pending_qty - qty
      found["last_action"] = {
        "type": action,
        "qty": qty,
        "rack_code": rack_code or "",
        "processed_at": datetime.datetime.now().isoformat(),
      }
      found.setdefault("logs", []).append(found["last_action"])
      day["updated_at"] = datetime.datetime.now().isoformat()
      payload["days"][date_text] = day
      self._write(payload)
      return json.loads(json.dumps(day, ensure_ascii=False))


_mops_lookup_engine = None

def _parse_inbound_date(date_text: str) -> str:
  try:
    return datetime.date.fromisoformat((date_text or "").strip()).isoformat()
  except Exception:
    raise RuntimeError("날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식이어야 합니다.")

def _normalize_header(value) -> str:
  return re.sub(r"\s+", "", str(value or "")).strip().lower()

def _to_number(value) -> int:
  if value is None:
    return 0
  if isinstance(value, bool):
    return int(value)
  if isinstance(value, (int, float)):
    return int(round(float(value)))
  text_value = str(value).strip().replace(",", "")
  if not text_value:
    return 0
  try:
    return int(round(float(text_value)))
  except Exception:
    return 0

def _get_lookup_engines() -> List:
  global _mops_lookup_engine
  engines = []
  if _mops_lookup_engine is None:
    try:
      _mops_lookup_engine = create_engine(
        MOPS_DB_URL,
        pool_pre_ping=True,
        future=True,
        pool_recycle=1800,
        isolation_level="READ COMMITTED",
      )
    except Exception:
      _mops_lookup_engine = False
  if _mops_lookup_engine:
    engines.append(_mops_lookup_engine)
  if isinstance(provider, MySQLProvider):
    engines.append(provider.engine)
  return engines

def _lookup_item_codes_by_name(names: List[str]) -> Dict[str, str]:
  unique_names = sorted({(name or "").strip() for name in names if (name or "").strip()})
  if not unique_names:
    return {}
  sql = text("SELECT code, name FROM items WHERE name IN :names ORDER BY code ASC")
  sql = sql.bindparams(bindparam("names", expanding=True))
  mapping: Dict[str, str] = {}
  for engine in _get_lookup_engines():
    try:
      with engine.connect() as conn:
        for code, name in conn.execute(sql, {"names": unique_names}):
          name_text = (name or "").strip()
          code_text = (code or "").strip()
          if name_text and code_text and name_text not in mapping:
            mapping[name_text] = code_text
      if mapping:
        break
    except Exception:
      continue
  return mapping

def _parse_new_inbound_workbook(content_bytes: bytes) -> List[dict]:
  workbook = load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)
  sheet = workbook.active
  headers = {
    idx: _normalize_header(sheet.cell(row=2, column=idx).value)
    for idx in range(1, sheet.max_column + 1)
  }

  product_col = next((idx for idx, title in headers.items() if "품명" in title), None)
  detail_qty_col = next((idx for idx, title in headers.items() if "상세수량" in title), None)
  box_qty_col = next((idx for idx, title in headers.items() if "박스수" in title), None)

  if not product_col or not detail_qty_col or not box_qty_col:
    raise RuntimeError("엑셀 2행에서 품명, 상세수량, 박스수 컬럼을 찾을 수 없습니다.")

  rows: List[dict] = []
  names: List[str] = []
  for row_idx in range(3, sheet.max_row + 1):
    product_name = str(sheet.cell(row=row_idx, column=product_col).value or "").strip()
    if not product_name:
      continue
    box_qty = _to_number(sheet.cell(row=row_idx, column=box_qty_col).value)
    inbound_qty = _to_number(sheet.cell(row=row_idx, column=detail_qty_col).value)
    row_item = {
      "id": uuid.uuid4().hex,
      "sku_code": "",
      "product_name": product_name,
      "box_qty": box_qty,
      "inbound_qty": inbound_qty,
      "pending_qty": inbound_qty,
      "logs": [],
    }
    rows.append(row_item)
    names.append(product_name)

  workbook.close()

  code_map = _lookup_item_codes_by_name(names)
  for row_item in rows:
    row_item["sku_code"] = code_map.get(row_item["product_name"], "")
  return rows


# ---------- Pydantic Models for API ----------
class CellItemModel(BaseModel):
  sku: str
  name: str
  qty: int
  used_cm3: int
  location: str = ""

class CellModel(BaseModel):
  code: str
  bay: int
  level: int
  capacity_cm3: int
  total_cm3: int
  kinds: int
  items: List[CellItemModel]

class CellsResponse(BaseModel):
  row: str
  bays: int
  levels: int
  cells: List[CellModel]

class ItemsResponse(BaseModel):
  items: List[dict]

class ConfigResponse(BaseModel):
  rows: List[str]
  bays: int
  levels: int
  level_heights_cm: dict
  rate_low_max: int
  rate_normal_max: int

class InboundRequest(BaseModel):
  rack_code: str
  item_code: str
  qty: int

class OutboundRequest(BaseModel):
  rack_code: str
  item_code: str
  qty: int

class MoveRequest(BaseModel):
  from_rack: str
  to_rack: str
  item_code: str
  qty: int

class SetLocationRequest(BaseModel):
  item_code: str
  location: str

class SearchResultModel(BaseModel):
  rack_code: str
  row: str
  bay: int
  level: int
  sku: str
  name: str
  qty: int
  capacity_cm3: int
  used_cm3: int
  rate: int
  location: str = ""

class SearchResultsResponse(BaseModel):
  q: str
  results: List[SearchResultModel]

class NewInboundImportRequest(BaseModel):
  date: str
  filename: str = ""
  content_base64: str

class NewInboundProcessRequest(BaseModel):
  date: str
  entry_id: str
  action: str
  qty: int
  rack_code: str = ""

class NewInboundListResponse(BaseModel):
  date: str
  updated_at: str = ""
  source_name: str = ""
  items: List[dict]

# ---------- FastAPI ----------
app = FastAPI(title="Warehouse PWA API", version="1.3.0")
app.add_middleware(
  CORSMiddleware,
  allow_origins=["*"],
  allow_credentials=False,
  allow_methods=["*"],
  allow_headers=["*"],
)

app.mount("/web", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "..", "web"), html=True), name="web")

@app.get("/")
def root():
  return RedirectResponse(url="/web/")

def make_provider() -> DataProvider:
  try:
    return MySQLProvider(DB_URL)
  except Exception as e:
    print("[WARN] MySQL 연동 실패, 더미 데이터로 전환:", e)
    return FakeProvider()

provider: DataProvider = make_provider()
new_inbound_store = NewInboundStore(NEW_INBOUND_STORE_PATH)

@app.get("/api/config", response_model=ConfigResponse)
def get_config():
  bays, levels = provider.layout()
  return ConfigResponse(
    rows=provider.rows(),
    bays=bays, levels=levels,
    level_heights_cm=LEVEL_HEIGHT_CM,
    rate_low_max=RATE_LOW_MAX,
    rate_normal_max=RATE_NORMAL_MAX,
  )

@app.get("/api/cells", response_model=CellsResponse)
def get_cells(row: str = ROW_IDS[0]):
  if row not in provider.rows():
    raise HTTPException(404, detail="없는 행(row)입니다.")
  bays, levels = provider.layout_for_row(row)  # 행별 레이아웃 사용
  data = provider.fetch_cells_for_row(row)
  out: List[CellModel] = []
  for (bay, level), ci in sorted(data.items(), key=lambda x: (x[0][0], x[0][1])):
    out.append(CellModel(
      code=ci.code, bay=bay, level=level,
      capacity_cm3=ci.capacity, total_cm3=ci.total_cm3,
      kinds=ci.kinds,
      items=[CellItemModel(sku=s, name=n, qty=q, used_cm3=u, location=l) for (s,n,q,u,l) in ci.items]
    ))
  return CellsResponse(row=row, bays=bays, levels=levels, cells=out)

@app.get("/api/movements")
def get_movements(limit: int = 30):
  rows = provider.recent_movements(limit=limit)
  return {"movements": [{"time":t, "type":typ, "sku":sku, "qty":qty, "path":path} for (t,typ,sku,qty,path) in rows]}

@app.get("/api/items", response_model=ItemsResponse)
def get_items(q: str = "", limit: int = 500):
  rows = provider.list_items(q, limit=limit)
  items = []
  for code, name, w, l, h, loc in rows:
    unit = int(max(0.0,w)*max(0.0,l)*max(0.0,h))
    items.append({"code":code, "name":name, "w":w, "l":l, "h":h, "unit_cm3":unit, "location":loc})
  return ItemsResponse(items=items)

@app.get("/api/items_with_stock", response_model=ItemsResponse)
def get_items_with_stock(q: str = "", limit: int = 300):
  rows = provider.list_items(q, limit=limit)
  codes = [code for (code, _name, _w, _l, _h, _loc) in rows]
  stock_map = {}
  # provider에 메서드가 있으면 사용
  if hasattr(provider, "latest_stock_for_codes"):
    stock_map = provider.latest_stock_for_codes(codes) or {}

  items = []
  for code, name, w, l, h, loc in rows:
    unit = int(max(0.0,w)*max(0.0,l)*max(0.0,h))
    items.append({
      "code": code,
      "name": name,
      "w": w, "l": l, "h": h,
      "unit_cm3": unit,
      "location": loc,
      "stock_qty": int(stock_map.get(code, 0)),  # ★ 최신 재고
    })
  return ItemsResponse(items=items)

@app.get("/api/search_racks", response_model=SearchResultsResponse)
def api_search_racks(q: str = "", limit: int = 500):
  results = provider.search_racks(q, limit=limit)
  return SearchResultsResponse(q=q, results=results)

@app.get("/api/new_inbound_list", response_model=NewInboundListResponse)
def get_new_inbound_list(date: str):
  try:
    date_text = _parse_inbound_date(date)
    return new_inbound_store.get_day(date_text)
  except Exception as e:
    raise HTTPException(400, detail=str(e))

@app.post("/api/new_inbound_list/import", response_model=NewInboundListResponse)
def post_new_inbound_list_import(req: NewInboundImportRequest):
  try:
    date_text = _parse_inbound_date(req.date)
    content_bytes = base64.b64decode(req.content_base64)
    items = _parse_new_inbound_workbook(content_bytes)
    return new_inbound_store.replace_day(date_text, items, source_name=req.filename)
  except Exception as e:
    raise HTTPException(400, detail=str(e))

@app.post("/api/new_inbound_list/process")
def post_new_inbound_list_process(req: NewInboundProcessRequest):
  try:
    date_text = _parse_inbound_date(req.date)
    action = (req.action or "").strip().lower()
    if action not in ("display", "inbound"):
      raise RuntimeError("지원하지 않는 처리 방식입니다.")

    current_day = new_inbound_store.get_day(date_text)
    item = next((entry for entry in current_day.get("items", []) if str(entry.get("id") or "") == str(req.entry_id)), None)
    if not item:
      raise RuntimeError("선택한 항목을 찾을 수 없습니다.")

    sku_code = str(item.get("sku_code") or "").strip()
    if not sku_code:
      raise RuntimeError("상품코드가 없는 항목은 처리할 수 없습니다.")

    updated_row = None
    rack_code = ""
    if action == "inbound":
      rack_code = (req.rack_code or "").strip()
      if not rack_code:
        raise RuntimeError("입고 처리에는 위치 선택이 필요합니다.")
      rack_norm = to_canonical_code(rack_code)
      provider.inbound(rack_norm, sku_code, req.qty)
      parsed = parse_rack_code(rack_code)
      updated_row = parsed[0] if parsed else None

    updated_day = new_inbound_store.process_item(date_text, req.entry_id, req.qty, action, rack_code=rack_code)
    response = {"ok": True, "date": date_text, "list": updated_day}
    if updated_row:
      response["row"] = updated_row
      response["cells"] = model_to_dict(get_cells(row=updated_row))
    return response
  except Exception as e:
    raise HTTPException(400, detail=str(e))

@app.post("/api/inbound")
def post_inbound(req: InboundRequest):
  try:
    code_raw  = (req.rack_code or "").strip()
    code_norm = to_canonical_code(code_raw)
    provider.inbound(code_norm, req.item_code, req.qty)
  except Exception as e:
    raise HTTPException(400, detail=str(e))
  parsed = parse_rack_code(code_raw)
  if not parsed:
    raise HTTPException(400, detail="잘못된 위치 코드")
  row, _, _ = parsed
  cells = get_cells(row=row)
  return {"ok": True, "row": row, "cells": model_to_dict(cells)}

@app.post("/api/outbound")
def post_outbound(req: OutboundRequest):
  try:
    code_raw  = (req.rack_code or "").strip()
    code_norm = to_canonical_code(code_raw)
    provider.outbound(code_norm, req.item_code, req.qty)
  except Exception as e:
    raise HTTPException(400, detail=str(e))
  parsed = parse_rack_code(code_raw)
  if not parsed:
    raise HTTPException(400, detail="잘못된 위치 코드")
  row, _, _ = parsed
  cells = get_cells(row=row)
  return {"ok": True, "row": row, "cells": model_to_dict(cells)}

@app.post("/api/move")
def post_move(req: MoveRequest):
  try:
    from_code  = (req.from_rack or "").strip()
    to_code    = (req.to_rack   or "").strip()
    from_norm  = to_canonical_code(from_code)
    to_norm    = to_canonical_code(to_code)
    provider.move(from_norm, to_norm, req.item_code, req.qty)
  except Exception as e:
    raise HTTPException(400, detail=str(e))
  pA = parse_rack_code(from_norm);  pB = parse_rack_code(to_norm)
  if not pA or not pB:
    raise HTTPException(400, detail="잘못된 위치 코드")
  rowA, _, _ = pA
  rowB, _, _ = pB
  return {
    "ok": True,
    "rows": {
      rowA: model_to_dict(get_cells(row=rowA)),
      rowB: model_to_dict(get_cells(row=rowB)),
    }
  }

@app.post("/api/set_location")
def post_set_location(req: SetLocationRequest):
  loc_raw = (req.location or "").strip()
  item_code = (req.item_code or "").strip()
  if not item_code:
    raise HTTPException(400, detail="item_code 누락")

  # ✅ 1) 정확히 "00"이면 중복검사 생략하고 그대로 저장
  if loc_raw == "00":
    try:
      _update_item_location(item_code, "00")
      return {"ok": True}
    except Exception as e:
      raise HTTPException(400, detail=f"위치 저장 실패: {e}")

  # 2) 그 외에는 기존처럼 중복검사 -> 저장
  try:
    if _check_location_in_use(loc_raw):
      raise HTTPException(409, detail="이미 사용중인 로케이션 코드로 등록에 실패했습니다.")
  except HTTPException:
    raise
  except Exception as e:
    raise HTTPException(400, detail=f"중복확인 실패: {e}")

  try:
    _update_item_location(item_code, loc_raw)
    return {"ok": True}
  except Exception as e:
    raise HTTPException(400, detail=f"위치 저장 실패: {e}")

def _update_item_location(item_code: str, loc: str):
  # provider가 MySQLProvider라고 가정
  if not isinstance(provider, MySQLProvider):
    raise RuntimeError("DB provider가 MySQL이 아닙니다.")
  col = provider._item_loc_col
  if not col:
    raise RuntimeError("items 테이블 또는 위치 컬럼을 찾을 수 없습니다.")

  sql = text(f"UPDATE items SET {col} = :loc WHERE code = :code")
  with provider.engine.begin() as conn:
    n = conn.execute(sql, {"loc": loc, "code": item_code}).rowcount
    if n <= 0:
      raise RuntimeError("해당 item_code를 찾을 수 없습니다.")

def _check_location_in_use(loc: str) -> bool:
  """이미 동일 로케이션이 다른 아이템에 설정되어 있는지 검사"""
  if not isinstance(provider, MySQLProvider):
    return False
  col = provider._item_loc_col
  if not col:
    return False

  # 정확히 "00"은 여기 오기 전에 처리되므로 여기선 그대로 검사
  sql = text(f"SELECT COUNT(*) FROM items WHERE {col} = :loc")
  with provider.engine.connect() as conn:
    cnt = int(conn.execute(sql, {"loc": loc}).scalar() or 0)
  return cnt > 0
